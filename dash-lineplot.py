################################################################
# The contents of this file are subject to the licenses listed below.
# You may not use this file except in compliance with these Licenses. 
# 
# Python, scipy, numpy, pandas, and other 'standard' modules are licensed 
# under the Python License: https://docs.python.org/3.7/license.html.  
#
# PySide: 'Qt for Python' is licensed under the LGPL3 license:
# https://www.gnu.org/licenses/lgpl-3.0.html
# 
# plotly/dash/visdcc is licensed under MIT https://community.plot.ly/t/pricing-and-license/9714
# https://en.wikipedia.org/wiki/MIT_License
#
################################################################

"""

This script reads an Excel config file and one or more of the following file types:
    * matlab file with data in 'DATA', variable names in 'NAM' and time base in 'TIME'
    * csv files with column names in top row
    * first sheet of an xlsx file with column names in top row
It the then proceeds to create and serve a Dash portal. 
The page served has several elements, all constructed from the 
information provided in the config file.

The config file has any number of sheets where each sheet defines
a different set of line graphs to be rendered on a separate tab 
(except for the header sheet, which defines the page header.)
Each graph sheet defines the height of the graphs, axes labels,
one x-value column name and any number of sets of y-value column names.
Each line has a number of attributes with default values if not supplied.
Each tab can be switched on/off for display purposes.
Each graph set can be exported to an html file.

The data file is read and a set of Dash data structures are formed
according to the Excel config file specifications.

In the present script the default config filename is './dash-config.xlsx'.
Any other filename can be provided on the commandline using the -f input flag.

Dash starts a Flask server at the specified port, so the browser must be 
pointing to the appropriate port number
localhost:port
This means that once the server is running, you can view the page with 
the PySide browser as used here, or in an external browser.

This module requires the following data in the current directory:
 * icons/logoSet2long.png
 * assets/bWLwgP.css

It will create folder 'graphs' for output. 

There are numerous Dash and Plotly resources on the Internet:
https://dash.plot.ly/integrating-dash
https://plot.ly/python/reference/
https://www.datacamp.com/community/tutorials/learn-build-dash-python
https://github.com/plotly/dash-recipes
https://github.com/plotly/dash-recipes/blob/master/multiple-hover-data.py
https://plot.ly/python/subplots/
https://towardsdatascience.com/creating-an-interactive-data-app-using-plotlys-dash-356428b4699c
https://dash.plot.ly/dash-core-components/tabs
https://dash.plot.ly/getting-started-part-2
https://plot.ly/python/range-slider/
https://plot.ly/python/click-events/

This script requires openpyxl, PySide2 (PyQt5 is loaded if PySide2 not available), numpy, 
pandas, plotly, dash, threading, openpyxl and some system modules.

To install dash when connected to the internet:
conda config --add channels conda-forge
conda search dash-daq --channel conda-forge
conda install dash
conda install dash-html-components
conda install dash-core-components
conda install dash-table
conda install dash-daq

This package could not be installed with 
    conda install visdcc
Conflicts between versions. Installing from the bz2 file worked however.

A recent off-line install required the following packages to be manually installed.
conda install dash-0.39.0-py_0.tar.bz2
conda install flask-compress-1.4.0-py_0.tar.bz2
conda install plotly-4.1.1-py_0.tar.bz2
conda install dash-html-components-0.14.0-py_0.tar.bz2
conda install dash-core-components-0.44.0-py_0.tar.bz2
conda install dash-table-3.6.0-py_0.tar.bz2
conda install dash-daq-0.1.4-py_0.tar.bz2
conda install plotly-orca-1.2.1-1.tar.bz2
conda install retrying-1.3.3-py37_1.tar.bz2
conda install dash-renderer-0.20.0-py_0.tar.bz2
conda install visdcc-0.0.40-pyh516909a_0.tar.bz2

Plotly packages seem to be here:  
https://anaconda.org/plotly  
https://anaconda.org/plotly/repo  
There are 17 packages, located under the package name, Files tab:
https://anaconda.org/plotly/plotly/files  
or   
https://anaconda.org/plotly/dash/files 

To use as a module in another application:

1) Import the DashLinePlot and DashPlotWindow classes from the module
            
2) In your code implement something like:

    # create new window
    self.dashWidget = DashPlotWindow(port)
    self.dashWidget.show()

    # do actual plotting
    useCallBacks = True
    plotConfig = './dash-config.xlsx'
    port = '8050' 
    dashlineplotter = DashLinePlot()
    dashlineplotter.runPlotter(port, plotConfig, useCallBacks)

Notes from https://dash.plot.ly/getting-started:
* The layout is composed of a tree of "components" like html.Div and dcc.Graph.
* The dash_html_components library has a component for every HTML tag. 
  Each html.xxx(children='yyy') component generates a <h1>yyy</h1> HTML element in your application.
* Not all components are pure HTML. The dash_core_components describe higher-level components that 
  are interactive and are generated with JavaScript, HTML, and CSS through the React.js library.
* Each component is described entirely through keyword attributes. 
  Dash is declarative: you will primarily describe your application through these attributes.
* The children property is special. By convention, it's always the first attribute which means that you can omit it: 
     html.xxx(children='yyy') is the same as html.xxx('yyy'). 
  Also, it can contain a string, a number, a single component, or a list of components.
* The fonts in the application can be set with a custom CSS stylesheet to modify the default styles of the elements. 
    external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css']
    app = dash.Dash(__name__, external_stylesheets=external_stylesheets)

https://dash.plot.ly/dash-html-components
The dash layout is composed of a tree of "components" like html.Div and dcc.Graph.
The dash_html_components library contains a component class for every HTML tag as well as keyword arguments 
for all of the HTML arguments.

https://dash.plot.ly/dash-core-components
The dash_core_components includes a set of higher-level components like dropdowns, graphs, markdown blocks, and more.
Graph renders interactive data visualizations using the open source plotly.js JavaScript graphing library. 
Plotly.js supports over 35 chart types and renders charts in both vector-quality SVG and high-performance WebGL.
The figure argument in the dash_core_components.Graph component is the same figure argument that is used by plotly.py, 
Plotly's open source Python graphing library. Check out the plotly.py documentation and gallery to learn more.

Notes on callbacks https://dash.plot.ly/getting-started-part-2:

# https://dash.plot.ly/dash-core-components/tabs
# A Div component is a wrapper for the <div> HTML5 element.
Div(
    [
        # The Tabs component hold a collection of Tab components.
        Tabs
        (
            # children (list of a list of or a singular dash component, string or numbers | a list of or a singular dash component, 
            # string or number; optional): Array that holds Tab components
            children=
            [
                # The Tab component controls the style and value of the individual tab 
                # id (string; optional): The ID of this component, used to identify dash components in callbacks. 
                #                        The ID needs to be unique across all of the components in an app.
                # label (string; optional): The tab's label
                # value (string; optional): Value for determining which Tab is currently selected
                #
                # Possible properties of Tab# ['children', 'id', 'label', 'value', 'disabled', 'disabled_style', 'disabled_className', 'className', 'selected_className', 'style', 'selected_style', 'loading_state']
                
                Tab(id='RelativePosition', label='RelativePosition', value='Tab 0'), 
                Tab(id='Velocity', label='Velocity', value='Tab 1'), 
                Tab(id='MissilePosition', label='MissilePosition', value='Tab 2'), 
                Tab(id='gimbalFromxls', label='gimbalFromxls', value='Tab 3')
            ], 
            
            # id (string; optional): The ID of this component, used to identify dash components in callbacks. 
            # The ID needs to be unique across all of the components in an app.
            id='tabs', 
            
            # value (string; optional): The value of the currently selected Tab
            value='Tab 0'
        ), 
        
        # id (string; optional): The ID of this component, used to identify dash components in callbacks. 
                                 The ID needs to be unique across all of the components in an app.
        Div(id='tabs-content')
    ]
    )
    
"""
__version__= '$Revision: 4633 $'
__author__='CJ & MS Willers'

import sys, os
import threading
import pandas as pd
import openpyxl as oxl
import numpy as np
import datetime   
import itertools 

# PySide2 is preferred based on licensing restrictions of PyQt5
try:
    __import__('PySide2')
    from PySide2 import QtWidgets
    import PySide2.QtCore as QtCore
    from PySide2 import QtWebEngineWidgets
except ImportError:
    try:
        __import__("PyQt5")
        from PyQt5 import QtWidgets
        import PyQt5.QtCore as QtCore
        from PyQt5 import QtWebEngineWidgets
    except ImportError:
        print("This script requires Python 3 with either PySide2 or PyQt5")
        exit(-1)
            
import dash
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output, State
from plotly import subplots
import visdcc

external_stylesheets = ['assets/bWLwgP.css']

pd.set_option('display.max_rows', 500)

# https://stackoverflow.com/questions/55596932/how-can-i-include-assets-of-a-dash-app-into-an-exe-file-created-with-pyinstaller
# when packaging the app with pyInstaller the assets folder is not included correctly
# defining resource_path as below and using
#     dash.Dash(__name__, assets_folder=resource_path('assets'))
# solves the problem 
def resource_path(relative_path):

# get absolute path to resource
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

################################################################
class DashLinePlot():

    def __init__(self):
        """
        Initialise class variables.
        """
        self.useCallbacks = True

        # storage for last 2 clicked point all graphs
        self.clickedData = {}

    ##########################################
    def generateFeedbackBoxes(self, id, isMarkers):
        """
        Builds the div with the click and rectangle tool feedback boxes

        Args:
            | id (string): id string.


        Returns:
            | thisDivList (list): list of html Divs.


        """

        # style definition for the data display boxes
        boxStyle = {
            'border': 'thin lightgrey solid',
            'overflowX': 'scroll'
        }

        # Divs for click data and rectangle tool data feedback
        # https://dash.plot.ly/interactive-graphing
        # https://dash.plot.ly/dash-html-components/pre

        clickDiv = html.Div(
                        [
                            dcc.Markdown(""" **Click Data** """), 
                            html.Pre(id='click-'+ id, style=boxStyle),
                        ], 
                        className='four columns'
                    )

        rectangleDiv =  html.Div(
                            [
                                dcc.Markdown(""" **Rectangle Tool Selection Data** """),
                                html.Pre(id='select-'+ id, style=boxStyle),
                            ], 
                            className='four columns'
                        )

        if isMarkers:
            return html.Div(className='row', children=[ clickDiv, rectangleDiv ])
        else:
            return html.Div(className='row', children=[ clickDiv ])
                

    def graphToDisk(self, figdict, fbasename):
        """
        Save the figure to disk as html

        Args:
            | figdict (dict): figure data
            | fbasename (string): file base name.

        Returns:
            | None.


        """
        # Save the figure to disk as html
        import plotly.offline as offline
        offline.plot(figdict,
            auto_open=False, 
            output_type='file', filename=f'{fbasename}.html', validate=False)


    ##########################################
    def makeGraphSet(self, dft, graph, reqStart = 0, reqEnd = 0):
        """
        Builds the set of graphs on this tab (requested from one sheet in xls) 

        Args:
            | dft (pd.dataframe): info for this graph set.
            | graph (string): graph set name, i.e. text following "graph-" in the sheet name.
            | reqStart (double): starting x-value, default the beginning.
            | reqEnd (double): ending x-value, default the end. 

        Returns:
            | thisDivList (list): list of html Divs.
            | grList (list): list of the symbolic names of all graphs in this set.
            | xmin (double): minimum x value
            | xmax (double): maximum x value

        """
        #  colors
        backgroundColor = 'aliceblue'
        gridColour = 'lightgrey'

        # get the header info from the header sheet in the config file
        pagetop = dfPlotterHeader.loc['PageTop','Value'] if 'PageTop' in dfPlotterHeader.index else ''
        pagebottom = self.dateCreated + ' ' + dfPlotterHeader.loc['PageBottom','Value'] if 'PageBottom' in dfPlotterHeader.index else ''
        
        # create graphs output folder if not exist
        grDir = './graphs'
        if not os.path.exists(grDir):
            os.mkdir(grDir)

        # get subplot bolean from the input
        # handle all graphs separately (default) or as subplots
        useSubplots = False
        if 'UseSubplots' in dft.index:
            if not np.isnan(dft[(dft['Variable']=='UseSubplots')]['Value'].values[0]):
                useSubplots = dft[(dft['Variable']=='UseSubplots')]['Value'].values[0]
                
        # graphs to disk requested?
        toDisk = True
        if 'ToDisk' in dft.index:
            if not np.isnan(dft[(dft['Variable']=='ToDisk')]['Value'].values[0]):
                toDisk = dft[(dft['Variable']=='ToDisk')]['Value'].values[0]

        # list of all graph names created here [passed back to calling function]
        # these names are the id of a Graph Div on the page, used in callback functions to update the figure
        grList = []

        # list of all the line entries for this graph set
        #  before building the page, all lines are first created and stored here
        graphData = []
     
        # get the filename for this graph to get to the data in the dataframe
        dfilename = dft[(dft['Variable']=='Datafile')]['Value'].values[0]
        df = self.datafiles[dfilename]

        # ------- x data preparation

        # 1) get the name of the x parameter
        xVarName = dft[dft['Variable']=='xValue']['Value'].values[0]

        # check requested x-range input validity and slice as requested
        if reqStart < df[xVarName].values[0]:
            reqStart = df[xVarName].values[0]
        if reqEnd > df[xVarName].values[-1]:
            reqEnd = df[xVarName].values[-1]
        if reqEnd <= reqStart:
            reqEnd = df[xVarName].values[-1]
        df = df[(df[xVarName] >= reqStart) & (df[xVarName] <= reqEnd)]

        #  2) get the graph set x hover text format from config
        hfmt_x = '.4f' 
        if isinstance(dft[dft['Variable']=='xLabel']['Format'].values[0], str):
            hfmt_x = dft[dft['Variable']=='xLabel']['Format'].values[0]

        # 3) apply the required scale and offset    
        if not np.isnan(dft[(dft['Variable']=='xValue')]['Scale'][0]):
            xscale = float(dft[(dft['Variable']=='xValue')]['Scale'][0])
        else:
            xscale = 1.0
            
        if not np.isnan(dft[(dft['Variable']=='xValue')]['Offset'][0]):
            xoffset = float(dft[(dft['Variable']=='xValue')]['Offset'][0])
        else:
            xoffset = 0.

        xData = df[xVarName] * xscale + xoffset

        # 4) slider marks dictionary based on set events in the data
        xmin = xData.min()
        xmax = xData.max()
        xsteps = 11
        sliderMarks={str(t): f'{t:.4f}s' for t in np.linspace(xmin,xmax,xsteps,endpoint=True)}

        # 5) step size of the x-axis slider
        xSliderStep = np.nan
        if 'xSliderStep' in dft['Variable']:
            xSliderStep =  dft[dft['Variable']=='xSliderStep']['Value'].values[0]           
        if np.isnan(xSliderStep):
            xSliderStep = round((xmax - xmin) / len(xData), 3)  

        #  6) all graphs on one page or tab have the same x label 
        xLabel = dft.loc['xLabel','Value']

        # ------- y data preparation

        # list of yValue values from config, i.e. the name of each variable to plot 
        yVariableList = []

        # build the traces for all required variables in this graph set
        for index, row in dft[(dft['Variable']=='yValue')].iterrows():

            # add to yValue line list
            yVariableList.append(index)

            # y scale
            if 'Scale' in row:
                if not np.isnan(row['Scale']):
                    yscale = row['Scale']
                else:
                    yscale = 1.0

            # y offset
            if 'Offset' in row:
                if not np.isnan(row['Offset']):
                    yoffset = row['Offset']
                else:
                    yoffset = 0.
                    
            # each line in each graph must be a dict as follows:
            plotMode = 'lines'
            if 'Mode' in row:
                #  a sting and not empty
                if isinstance(row['Mode'], str) and not row['Mode'] == "":
                    plotMode = row['Mode']

            opacity = 0
            if 'MarkerOpacity' in row:
                if not np.isnan(row['MarkerOpacity']):
                    opacity = row['MarkerOpacity']
                    
            markerDict = { 'opacity': opacity }

            dLines = {
                'x':xData,
                'y':df[row['Value']] * yscale + yoffset,
                'line':{},
                'mode': plotMode,
                'marker': markerDict,   # we do not want markers but need them for the rectangle tool to appear
            }

            # fill in non-default values
            if not np.isnan(row['Linewidth']):
                dLines['line']['width'] = row['Linewidth']

            if isinstance(row['Colour'], str):
                dLines['line']['color'] = row['Colour']

            if isinstance(row['Dash'], str):
                dLines['line']['dash'] = row['Dash']

            if isinstance(row['GraphType'], str):
                dLines['type'] = row['GraphType']

            if isinstance(row['LineLabel'], str):
                dLines['name'] = row['LineLabel']
            else:
                dLines['name'] = row['Value']

            dLines['showlegend'] = True

            # add this line to other lines in this graph
            graphData.append(dLines)

        # ------- html Div's preparation
        thisDivList = []

        # 1) Div header: append the header text at the top of the page
        thisDivList.append(
            html.Div([dcc.Markdown(id=f'headerMarkdown-{graph}',children=pagetop)]),
        )  

        # 2) Div top text: if supplied, append the sheet top text  
        if 'GraphTop' in dft.index:
            thisDivList.append(
                html.Div([dcc.Markdown(id=f'topMarkdown-{graph}',children=dft.loc['GraphTop','Value'])])
            )            

        # 3) Div x-axis slider 
        instruction = '**Click on current tab to refresh the x-axis slider and the graphs**'

        setName = graph
        thisDivList.append(
            html.Div([
                dcc.Markdown(id='header-xSlider-'+ setName,children=instruction),
                dcc.RangeSlider(
                    id='xSlider-'+ setName, min=xData.min(), max=xData.max(),  step=xSliderStep, 
                    value=[xData.min(), xData.max()],  
                    marks=sliderMarks, 
                    allowCross=False,
                    tooltip={'always_visible': False, 'placement': 'bottom'},  # use either the tooltip or the text display in next div
                    # updatemode='drag',   # default is mouseup
                    className='margin150'
                ),
                html.Div(
                    style={'marginTop':40, 'fontSize':12},
                    id='output-container-xSlider-'+ setName,
                    className='margin150'
                ),
                dcc.Input(id='minVal-'+ setName, type='number', min=0, step=xSliderStep, placeholder='type start value', className='margin150-2', style={'fontSize':12}),
                dcc.Input(id='maxVal-'+ setName, type='number', min=0, step=xSliderStep, placeholder='type end value', className='margin2', style={'fontSize':12}),
                html.Button(id='submit-button-'+ setName, type='submit', children='Submit', className='margin2'),
                html.Button('Reset slider', id='resetSlider-'+ setName, className='margin2'),
            ])
        )

        # 4) Graph and data feedback Divs

        # title rows
        titleRows = dft[(dft['Variable']=='Title')]

        #  graph titles
        grTitles = [row['Value'] for index, row in titleRows.iterrows()]

        # number of graph sets 
        numGraphSets = len(titleRows.index)

        # subplot environment setup to be done before running through the data collection
        if useSubplots:

            # for subplots we have only one Graph Div
            # use the graph set name without any added numbers
            grList.append(graph)

            # row heights
            grHeight = dft.loc['Height','Value']
            rowHeights = [grHeight] * numGraphSets

            # generate the subplot figure
            figdict = subplots.make_subplots(rows=numGraphSets, cols=1,
                                        shared_xaxes=True, shared_yaxes=False,
                                        vertical_spacing=0.075,
                                        row_heights=rowHeights,
                                        x_title=xLabel, 
                                        subplot_titles=grTitles
                                        )
            figdict.update_xaxes(hoverformat=hfmt_x, 
                                gridcolor=gridColour   # default is white
                                )  
            figdict.update_layout(hovermode='x', 
                                    plot_bgcolor=backgroundColor, 
                                    font=dict(size=10))   # setting the font size of all y-axes labels and legends


        # subplot counter used to pack the graph data to the figdict
        subNum = 0

        #  collect the data for the graphs by running through each set
        for index, row in titleRows.iterrows():

            # increase graph set counter, starting from one
            subNum = subNum + 1

            # get the set number as a string
            setStr = str(index).split('#')[1]

            #  current graph title and ylabel for the plot
            grTitle = row['Value'] 
            yLabel = dft.loc['yLabel#'+setStr,'Value']

            #  graph set y hover text format 
            hfmt_y = '.4f' 
            if isinstance(dft.loc['yLabel#'+setStr,'Format'], str):
                hfmt_y = dft.loc['yLabel#'+setStr,'Format']

            #  determine if the rectangle tool is present
            #  this will be the case if in any line is using markers 
            isMarkers = False

            # pack the graph data in
            #  * either a list to be used in the Graph Div
            #  * or in the relevant subplot 
            thisGraphData = []

            # run through all variables in the trace set
            for traceNum, value in enumerate(yVariableList):
                # identity the specific trace
                if 'yValue#'+setStr in value:

                    # add to plot set    
                    if useSubplots:
                        figdict.append_trace(graphData[traceNum], subNum, 1) 
                        figdict.update_yaxes(
                            hoverformat=hfmt_y, 
                            title=yLabel, 
                            gridcolor=gridColour,
                            row = subNum, col = 1) 
                    else:                 
                        thisGraphData.append(graphData[traceNum])
                        
                    # check for usage of markers
                    # at least one trace with markers will trigger the rectangle tool
                    # with associated Rectangle Tool Selection Data box
                    if 'markers' in graphData[traceNum]['mode']:
                        isMarkers = True

            # Not using subplots we create a Graph Div for each set 
            if not useSubplots:

                # create dictionary with the layout and data 
                figdict = {'layout':{'title': grTitle,
                                    'xaxis':{'title': xLabel, 'hoverformat': hfmt_x},
                                    'yaxis':{'title': yLabel, 'hoverformat': hfmt_y},	
                                    'clickmode': 'event+select',
                                    'hovermode': 'x',           # set compare data on hover
                                    'plot_bgcolor': backgroundColor, 
                                    },
                            'data':thisGraphData}
           
                #  store the id of this set - to be used in callback function generation
                #  we mark all relevant Divs with this string
                grID = graph+setStr
                grList.append(grID)

                # Div with dcc.Graph using the figdict
                thisDivList.append(
                    html.Div
                    (
                        [
                            dcc.Graph
                            (
                                id=grID,
                                figure=figdict,
                                style={'height': str(dft.loc['Height','Value']),'padding':20},
                            )
                        ]
                    )
                )

                # Divs for click data and rectangle tool data feedback
                thisDivList.append(self.generateFeedbackBoxes(grID, isMarkers))
                
                if toDisk:
                    self.graphToDisk(figdict, f'{grDir}/{graph}#{setStr}')

        # only one Graph Div if all graphs are in subplots
        if useSubplots:

            figdict.update_layout(height=numGraphSets*grHeight)  

            # Div with dcc.Graph using the figdict
            thisDivList.append(
                html.Div
                (
                    [
                        dcc.Graph
                        (
                            id=graph, 
                            figure=figdict,
                        ),
                        visdcc.Run_js(id='hover-js')  # need this to get the hover data on all lines of all subplots simultaneously
                    ]
                )
            )

            # Divs for click data and rectangle tool data feedback          
            thisDivList.append(self.generateFeedbackBoxes(graph, isMarkers))

            if toDisk:
                self.graphToDisk(figdict, f'{grDir}/{graph}')

        # 5) Div bottom text: if supplied, append the sheet bottom text
        if 'GraphBottom' in dft.index:
            thisDivList.append(
                html.Div([dcc.Markdown(id=f'botMarkdown-{graph}',children=dft.loc['GraphBottom','Value'])])
            )

        # 6) Div page footer: append the footer text at the bottom of the graph
        thisDivList.append(
            html.Div([dcc.Markdown(id=f'footerMarkdown-{graph}',children=pagebottom)]),
        ) 

        # 7) Div with license logos
        import base64
        encoded_image = base64.b64encode(open('icons/logoSet2long.png', 'rb').read())
        thisDivList.append(
            html.Div([
                        html.Img(src='data:image/png;base64,{}'.format(encoded_image.decode()),
                        height=50)
                    ], 
                    style = {'text-align':'right'}
                    )
        )

        return thisDivList, grList, xData.min(), xData.max()

    ##########################################
    def prepareGraphs(self):
        """
        Controls preparation of all required graph sets

        Args:
            | None. 

        Returns:
            | None.

        """
        #  declare global to enable changing
        global divSets
        global graphTabs
        global graphList
        global sliderMinValues
        global sliderMaxValues

        # divSets to be used when constructing the page
        # each entry in this list is a different tab containing several graphs
        divSets = []

        # list with the names of the tabs on the page
        graphTabs = []

        #  List of all the unique graph names for which we need to register callback functions
        graphList = []

        # slider limits
        sliderMinValues = []
        sliderMaxValues = []
        
        # make a list of all possible graph tabs and graphs sets in dataframe dfg
        # to be used in generating all possible callbacks
        global allTabs, allTabUsedIdx
        allTabs = dfPlotterConfig['Graph'].unique()
        allTabUsedIdx = [-1] * len(allTabs)

        global allGraphs
        allGraphs = []

        # counter for active tabs
        tabIndex = 0

        # for each graph tab in the input data, i.e. each sheet starting with 'graph-'
        for i, graphTab in enumerate(allTabs):

            # extract info for this graph set
            dft = dfPlotterConfig[(dfPlotterConfig['Graph']==graphTab)]        
            
            # extract all graph names for this tab    
            titleRows = dft[(dft['Variable']=='Title')]
            for index, row in titleRows.iterrows():
                setStr = str(index).split('#')[1]
                grID = graphTab+setStr
                allGraphs.append(grID)
            
            # First check exclude flag
            toInclude = True
            if 'Include' in dft.index:
                if not np.isnan(dft[(dft['Variable']=='Include')]['Value'].values[0]):
                    toInclude = dft[(dft['Variable']=='Include')]['Value'].values[0]
            
            # collect data and build the data for the sheet
            if toInclude:
                allTabUsedIdx[i] = tabIndex
                tabIndex = tabIndex + 1

                divSet, grList, xmin, xmax = self.makeGraphSet(dft, graphTab)

                divSets.append(divSet)
                graphList.append(grList)
                sliderMinValues.append(xmin)
                sliderMaxValues.append(xmax)
                graphTabs.append(graphTab.split('-')[1])

    ##########################################
    def makePage(self):
        """
        Create the page to be displayed in the browser

        Args:
            | None.

        Returns:
            | page (html Div): created page.

        """

        # each entry in this list is a different tab containing several graphs
        lsttabs = []

        # create each tab in the page
        # lsttabs will now have for each active tab: label & value
        #    e.g. [Tab(id='RelativePosition', label='RelativePosition', value='Tab 0'), 
        #          Tab(id='Velocity', label='Velocity', value='Tab 1'), 
        #          Tab(id='xyPlot', label='xyPlot', value='Tab 2'), 
        #          Tab(id='MissilePosition', label='MissilePosition', value='Tab 3'), 
        #          Tab(id='Attitude', label='Attitude', value='Tab 4'), 
        #          Tab(id='gimbalFromxls', label='gimbalFromxls', value='Tab 5')]
        for tabNum, tabSet in enumerate(divSets):

            tabLabel = graphTabs[tabNum]

            # if we use callbacks only the tab is created, i.e. no data added
            # the graphs are only added to the tab when the user clicks on the tab
            # callbacks are preferred for large datasets 
            if self.useCallbacks:
                lsttabs.append(
                    dcc.Tab(value='Tab ' + str(tabNum), label=tabLabel, id=tabLabel))

            # no callbacks, add the graphs on all tabs
            else:
                lsttabs.append(
                    dcc.Tab(value='Tab ' + str(tabNum), label=tabLabel, id=tabLabel, children=[
                        *tabSet,
                    ]))

        # create the page to be rendered in the browser, using all active tabs as requested via config
        page = html.Div(
        [
            dcc.Tabs(
                id='tabs', 
                value='Tab 0', 
                children=[
                # following is a list of all tabs with their content
                *lsttabs,
                ]
            ),

            html.Div(id='tabs-content'),
        ]
        )

        # the page now has for example:
        # Div([Tabs(
        #       children=[
        #         Tab(label='RelativePosition', value='Tab 0'), 
        #         Tab(label='Velocity', value='Tab 1'), 
        #         Tab(label='MissilePosition', value='Tab 2'), 
        #         Tab(label='gimbalFromxls', value='Tab 3')], 
        #       id='tabs', value='Tab 0'
        #      ),
        #      Div(id='tabs-content')
        #    ])
        # 
        #  If we are using callbacks to populate the tabs, we pass as input: value of the clicked tab, i.e. 
        #    Input(component_id='tabs', component_property='value'), value is one of Tab 0, Tab 1, Tab 2 or Tab 3
        #  and the page gets the output from the callback:
        #    Output(component_id='tabs-content', component_property='children'), passing back the graph set for this tab divSets[tabNum]
        #  This set includes all headers, graphs & footers for the selected tab

        return page

    ##########################################
    def loadConfig(self, configfile):
        """
        Loads the graph configuration from the excel file
            
        Args:
            | configfile (string): Excel filename for file that defines the plots. 

        Returns:
            | None.

        """

        # read the config file
        cxls = pd.ExcelFile(configfile)

        # header dataframe, i.e the data on the 'header' tab in the xlsx file 
        global dfPlotterHeader
        dfPlotterHeader = pd.read_excel(cxls, 'header')
        dfPlotterHeader = dfPlotterHeader.set_index('Variable')
        masterDataFile =  dfPlotterHeader.loc['Datafile','Value']

        # get a list of graph sheetnames (ignore the header sheet)
        cwb =  oxl.load_workbook(configfile)
        sheetnames = [sn for sn in cwb.sheetnames if 'graph' in sn]

        # dataframe to contain ALL the sheets' info
        global dfPlotterConfig
        dfPlotterConfig = pd.DataFrame()

        for shtnum,sheetname in enumerate(sheetnames):
            dft = pd.read_excel(cxls, sheetname)
            # add info to identify the lines associated with this sheet
            dft['Graph'] = sheetname
            dft['ShtNum'] = shtnum
            dft['Index'] = dft['Variable']

            # Check the file to be used and
            # determine the number of graphs on this tab
            i = 0
            theSet = -1
            for index,row in dft.iterrows():
                if 'Datafile' in row['Variable']:
                    if dft.loc[index,'Value'] == 'master':
                        dft.loc[index,'Value'] = masterDataFile
                if 'Title' in row['Variable']:
                    theSet = theSet + 1
                    dft.loc[index,'Index'] = f"{row['Variable']}#{theSet:03d}"
                if 'yLabel' in row['Variable']:
                    dft.loc[index,'Index'] = f"{row['Variable']}#{theSet:03d}"
                    i = 0
                if 'yValue' in row['Variable']:
                    dft.loc[index,'Index'] = f"{row['Variable']}#{theSet:03d}-{i:03d}"
                    i = i + 1

            # make 'Index' column the index
            dft = dft.set_index('Index')
            # append this sheet to the master data frame
            dfPlotterConfig = dfPlotterConfig.append(dft)

###########################################################################
    def readdatafile(self, filename):
        """Read a comma or space separated data file into a dataframe.

        OSSIM data files can use comma or space separated data files.
        These files normally have one comma or one or more space separators. 
        The header line for space separated files start with a percentage to allow
        loading of the file with Matlab. There might be a space between the % en the 
        name of the first column, e.g. '%time' or '% time'. Using pandas.read_csv() 
        will not work if a space is present between the % and the column name.

        This function tries to read the different styles of files into a 
        Pandas dataframe, removing the percentage and any spaces before the first column 
        heading (rest as in data file).
            
        The number of header lines may vary, discard all lines that starts with %
        using the last as header.
            
        Args:
            | filename (string): csv filename. 

        Returns:
            | dfData (pandas.DataFrame): dataframe with loaded data.
        
        """
        # empty dataframe
        dfData = None
            
        # determine if there are lines to skip before header line
        skiprows = 0
        headerDone = False
        with open(filename,'r') as fin:
            while not headerDone:
                line=fin.readline()
                if '%' in line:
                    skiprows = skiprows + 1
                else:
                    headerDone = True
                    
        # leave the last line as header  
        skiprows = skiprows - 1
    
        # identify the file type
        with open(filename,'r') as fin:
            line = fin.readline()
            if len(line) > 0:
                matlab = True if '%' in line else False
                matlabspace = True if ' ' == line[1] else False
                comma = True if ',' in line else False
            else:
                print('File {} has no contents, returning None'.format(filename))
                return None
        
        # load the data if matlab or space separated

        if matlab or '.plt' in filename:
            df = pd.read_csv(filename, sep='\s+',engine='python',header=0,skiprows=skiprows)
            
            # the leading '% ' in header messes up the column headings, fix:
            if matlabspace:
                dfData = df[df.columns[:-1]]
                dfData.columns = df.columns[1:]
            else:
                dfData = df

            if '%Time' in dfData.columns.values[0]:
                dfData['Time'] = dfData['%Time']
                dfData.drop(['%Time'], axis=1,inplace=True)
            if '%CurrentSimTime' in dfData.columns.values[0]:
                dfData['CurrentSimTime'] = dfData['%CurrentSimTime']
                dfData.drop(['%CurrentSimTime'], axis=1,inplace=True)
            if '%t' in dfData.columns:
                dfData['t'] = dfData['%t']
                dfData.drop(['%t'], axis=1,inplace=True)
                
        # load comma separated data
        if comma or '.csv' in filename:
            dfData = pd.read_csv(filename, sep=',',header=0)

        # load spectral data
        if '.scd' in filename or '.spc' in filename:
            dfData = pd.read_csv(filename, delimiter='\s+',header=None)
            if dfData.shape[1] == 3:
                dfData.columns=['wavelen','wavenum','trans']
            else:
                dfData.columns=['wavelen','wavenum','emis','trans','refl']

        return dfData

    ##########################################
    def loadData(self):
        """
        Load all the data from all files supplied

        Args:
            | None. 

        Returns:
            | success (bolean): True if the file load was successful.
            
        """
    
        # get data filenames from all sheets
        datafilenames = dfPlotterConfig[(dfPlotterConfig['Variable']=='Datafile')]['Value'].unique()

        self.datafiles = {}
        self.dateCreated = str(datetime.date.today())

        # run through all unique file names

        success = True
        for datafilename in datafilenames:

            if os.path.isfile(datafilename):

                # determine what type of file is this by looking at the file extension
                extension = os.path.splitext(datafilename)[1]

                # matlab format files
                # note that here we rely on the Denel GTV matlab file which has 
                #  * the data stored in 'DATA'
                #  * the data column names in 'NAM'
                #  * the time variable is called 'TIME'
                # if other applications need matlab file capability this must be generalised
                if 'mat' in extension:

                    # load the gtv telemetry data in matlab format file
                    # scipy reads in structures as structured numpy arrays of dtype object
                    # returns a dictionary with variable names as keys, and loaded matrices as values.
                    from scipy.io import loadmat
                    dataMat = loadmat(datafilename)

                    # create the dataframe
                    self.datafiles[datafilename] = pd.DataFrame(dataMat['DATA'], columns=dataMat['NAM'])

                    # set beginning of data set as time zero
                    self.datafiles[datafilename]['TIME'] = self.datafiles[datafilename]['TIME'] - self.datafiles[datafilename]['TIME'][0]
                    
                    # get date 
                    self.dateCreated = dataMat['Date_Created'][0]

                # Excel data files
                # top row is data column names
                # Only the first sheet is loaded
                # To be generalised to specify the sheet from the config file
                elif 'xls' in extension:
                    self.datafiles[datafilename] = pd.read_excel(datafilename, index_col=None)

                #  csv files
                #  top line is column names
                else:
                    self.datafiles[datafilename] = self.readdatafile(datafilename)
                    # pd.read_csv(datafilename, sep="\s+|,|;", index_col=None,engine='python')
    
            else:
                print(f'Data file {datafilename} for plotting not found, please provide a valid file name in the config file!\n ')
                success = False

        return success

    ##########################################
    #
    def run_dash(self, pageLayout,port):
        """
        Initiate the Dash server and serve the page

        Args:
            | pageLayout (dash layout): info the be served in Plotly data format
            | port (int): port number to be used

        Returns:
            | None.
            
        """
        # start a dash app, which also starts a Flask server
        # it is important to set the name parameter of the Dash instance to the value __name__, 
        # so that Dash can correctly detect the location of any static assets inside an assets 
        # directory for this Dash app
        # this must be global to stay in scope in applications that use the plotter as a module
        global dashApp
        dashApp = dash.Dash(__name__, external_stylesheets=external_stylesheets, assets_folder=resource_path('assets'))

        # override security restrictions: allow the serving of local pages
        dashApp.css.config.serve_locally = True
        dashApp.scripts.config.serve_locally = True
        dashApp.layout = pageLayout

        # We have a dynamic layout, so we can ignore the exception
        dashApp.config['suppress_callback_exceptions']=True

        # generate all callback functions for all possible graph sets & tabs
        self.setupCallbacks()

        # run the server on the specified port
        # set debug mode to False, no hot reloading
        # From https://dash.plot.ly/devtools
        # dev_tools features are activated by default when you run the app with app.run_server(debug=True)
        # By default, Dash includes "hot-reloading". This means that Dash will automatically refresh your browser 
        # when you make a change in your Python or CSS code.
        dashApp.run_server(debug=False, port=port, use_reloader=False)

    def setupCallbacks(self):
        """
        Generate all callback functions required

        Args:
            | None.

        Returns:
            | None.
            
        """
        # prepare for hover labels accross shared axes
        #   * can only be done when doing subplots
        #   * used the tricks from here
        #        https://github.com/plotly/plotly.js/issues/2114#issuecomment-535259328
        #  The main tricks are
        #    1) dynamically get plot names
        #    2) use visdcc.Runjs to reload the javasript on graph change, to reattach the event handler to 
        #       the re-created plot. 
        # 
        # generate javascript strings to run in the render callback
        # the plotid in this javascript string will be replaced with the graph id's
        JS_STR_template = '''

            var plotid = 'theplotname'
            var plot = document.getElementById(plotid)

            plot.on(
            'plotly_hover',
            function (eventdata) {
                Plotly.Fx.hover(
                plotid,
                { xval: eventdata.xvals[0] },
                Object.keys(plot._fullLayout._plots) // ["xy", "xy2", ...]
                );
            });
            '''
            
        jsString = []
        for gr in allTabs:
            theGraph = str(gr)
            gr_js = JS_STR_template.replace('theplotname',theGraph)
            jsString.append(gr_js)

        # ----------------------------------------------------------------------------------------------
        # now define all the callback functions:

        # construct the dependencies input list [not working yet!]
        # https://community.plot.ly/t/flexible-programmatic-variables-to-a-callback/7000
        # inputList = ['output-container-xSlider-'+str(i) for i in self.graphTabs]
        # INPUTS = [Input('tabs', 'value')]
        # INPUTS.extend([Input(i, 'children') for i in inputList])

        # callback used for rendering of tabs
        @dashApp.callback(
            [Output('tabs-content', 'children'),
            Output('hover-js', 'run'),  # <-- add this to get hover on all subplot traces
            ],
            [Input('tabs','value')]
            # INPUTS
        )
        def render_content(tab):
        # def render_content(tab, *args):
            tabNum = int(tab.split(' ')[1])
            # get the correct tab number for the js_str in complete list
            jsIndex = allTabUsedIdx.index(tabNum)
            js_str = jsString[jsIndex]
            return divSets[tabNum], js_str
    
        # generate data clicked and selected callback functions for all possible graphs in the config
        # i.e. subplots as well as individual graph sets
        # must be able to handle changed config input from the user
        for gr in itertools.chain(allTabs,allGraphs):
            theGraph = str(gr)

            # initialise the clicked data storage
            # current click index, click1, click2, range
            data = [1, [0,0], [0,0], [0,0]]
            self.clickedData[theGraph] = data

            @dashApp.callback(
                Output('click-'+theGraph, 'children'), # display box id and children
                [Input(theGraph, 'clickData')],   # graph id and clickdata
                [State(theGraph,'id')]
            )
            def display_click_data(clickData, id):
                msg = 'none clicked'
                if clickData:
                    # get clicked data
                    x = clickData['points'][0]['x']
                    y = clickData['points'][0]['y']

                    # Index of new click data
                    index = self.clickedData[id][0]

                    # store the new data here
                    self.clickedData[id][index][0] = x
                    self.clickedData[id][index][1] = y

                    # Calc the delta and set the index to be valid for next click
                    
                    indCur = index
                    if index == 1:
                        index = 2
                    else:
                        index = 1
                    indexPrev = index
                    self.clickedData[id][0] = index

                    # calc delta
                    self.clickedData[id][3][0] = abs(self.clickedData[id][indCur][0] - self.clickedData[id][indexPrev][0])
                    self.clickedData[id][3][1] = abs(self.clickedData[id][indCur][1] - self.clickedData[id][indexPrev][1])

                    msg =  (
                            f'Previous [x, y]: [{self.clickedData[id][indexPrev][0]:.6f}, {self.clickedData[id][indexPrev][1]:.6f}]\n'  
                            f'Current [x, y]: [{self.clickedData[id][indCur][0]:.6f}, {self.clickedData[id][indCur][1]:.6f}]\n'  
                            f'Range [x, y]: [{self.clickedData[id][3][0]:.6f}, {self.clickedData[id][3][1]:.6f}]' 
                    )

                return msg 

            @dashApp.callback(
                Output('select-'+theGraph, 'children'), # display box id and children
                [Input(theGraph, 'selectedData')]   # graph id and selectedData
            )
            def display_selected_data(selectedData):
                msg = 'none selected'

                if selectedData:

                    # for divs where we work with subplots, the number of the subplot is added to the
                    # x and y key. Get the keys programmatically.
                    rangeDict = selectedData['range']
                    keys = []
                    for key in rangeDict:
                        keys.append(key)

                    x1eft = rangeDict[keys[0]][0]
                    xright = rangeDict[keys[0]][1]
                    dx = abs(xright - x1eft)

                    ytop = rangeDict[keys[1]][1]
                    ybottom = rangeDict[keys[1]][0]
                    dy = abs(ybottom - ytop)
                    
                    msg = (
                        f'Top left [x, y]: [{x1eft:.6f}, {ytop:.6f}]\n'  
                        f'Bottom right [x, y]: [{xright:.6f}, {ybottom:.6f}]\n'  
                        f'Range in [x, y]: [{dx:.6f}, {dy:.6f}]' 
                    )

                return msg 

        # time slider callback for each tab - display selected values of the slider
        for gr in allTabs:
            theGraph = str(gr)

            @dashApp.callback(
                Output('output-container-xSlider-'+ theGraph, 'children'),
                [Input('xSlider-'+theGraph, 'value'),
                 Input('submit-button-'+theGraph, 'n_clicks'), 
                ],    
                [State('tabs', 'value'),
                 State('minVal-'+theGraph, 'value'), State('maxVal-'+theGraph, 'value'),
                ]             
            )
            def process_xSlider_data(value, nclicks, tab, mini, maxi):
                # tab number in the current page layout
                tabNum = int(tab.split(' ')[1])
                graphSetName = 'graph-'+graphTabs[tabNum]
                # select the graph data
                dft = dfPlotterConfig[(dfPlotterConfig['Graph']==graphSetName)] 
                # determine which input triggered the callback
                ctx = dash.callback_context
                clicked_id = ctx.triggered[0]['prop_id'].split('.')[0]
                # Get slider limits from input fields
                if 'submit' in clicked_id:
                    start = mini
                    if start < sliderMinValues[tabNum]:
                        start = sliderMinValues[tabNum]
                    end = maxi
                    if end > sliderMaxValues[tabNum]:
                        end = sliderMaxValues[tabNum] 
                    value[0] = start
                    value[1] = end

                # update the graph set
                global divSets
                divSets[tabNum], _, _, _ = self.makeGraphSet(dft, graphSetName, value[0], value[1]) 
                msg = f'Selected range [{value[0]:.6f}, {value[1]:.6f}]'
                return msg
            
            @dashApp.callback(
                [Output('xSlider-'+theGraph, 'value'), 
                 Output('minVal-'+theGraph, 'value'), 
                 Output('maxVal-'+theGraph, 'value'), 
                ],
                [Input('resetSlider-'+theGraph, 'n_clicks')],
                [State('tabs', 'value')] 
            )
            def reset_xSlider(nclicks, tab):
                tabNum = int(tab.split(' ')[1])
                low = sliderMinValues[tabNum]
                hi = sliderMaxValues[tabNum]
                limit = [low, hi]
                return limit, '', ''

    ##########################################
    def runPlotter(self, port, configfile, cback = True, flaskServerRunning=False):
        """
        main control plotter function

        Args:
            | configfile (string): Excel configuration file defining the graphs.
            | cback (bolean): use callbacks to populate the data on the tabs (default True)
                             (recommended for large data sets)
            | flaskServerRunning (bolean): entry state of the flask server (default False)

        Returns:
            | flaskServerRunning (bolean): running state of flask server at the end of this function.
        """

        # set callbacks flag as requested
        self.useCallbacks = cback

        sys.argv.append("--disable-web-security")

        self.loadConfig(configfile)

        # load all data to be available in the class 
        # all the data files, but only once into a dict with filename as key
        
        if self.loadData():           
            # prepare all required graph sets
            self.prepareGraphs()

        
            # now create the page we want to render
            pageLayout = self.makePage() 

            # The Python threading API defines two kinds of threads: daemons and non-daemons. 
            # A Python program is defined to end when all non-daemons are done. 
            # So, if your thread is infinite, which they often are, they will never be done and your program is hard to exit.
            # Make the thread a daemon, i.e. process running in the background, so that we can easily kill the program.
            # A daemon thread will shut down immediately when the program exits. One way to think about these definitions is to 
            # consider the daemon thread a thread that runs in the background without worrying about shutting it down.

            # the first entry to this function starts the thread as a daemon
            # this means that the user can close the dash window, change the configuration in the
            # setup file, open a new dash window, then only render the page with the updated information 
            # as implemented in the else section here.
            if not flaskServerRunning:
                sys.argv.append("--disable-web-security")
                threading.Thread(target=self.run_dash, args=(pageLayout,port), daemon=True).start()
                flaskServerRunning = True
            else:
                # serve new page
                dashApp.layout = pageLayout

        return flaskServerRunning       
    
##########################################
#
class WebViewer(QtWebEngineWidgets.QWebEngineView):
    """
    creates a web engine view widget

    """
    def __init__(self, parent, url):
        """
        Initialise the web browser widget

        Args:
            | parent (GUI element): the parent GUI element where this widget is included.
            | url (url):the url to be browsed

        Returns:
            | None.

        """
        super().__init__(parent)

        # ensure the complete view has the same style
        # if this is not present, the tabs as well as top and bottom markdown
        # have different style - only experienced when used as module 
        self.setStyleSheet(external_stylesheets[0])

        # create the page
        page = QtWebEngineWidgets.QWebEnginePage(self)
        self.setPage(page)
        self.setUrl(QtCore.QUrl(url))

##########################################
# 
class DashPlotWindow(QtWidgets.QMainWindow, QtWidgets.QWidget):    
    """
    creates a window to run the dash server in
    """                     

    def __init__(self, port, title):
        """
        Initialise the window

        Args:
            | port (int): the port to be used by the server.
            | title (string): window title

        Returns:
            | None.

        """
        super().__init__()
        self.setWindowTitle(title)
        self.setMinimumSize(640,640)
        
        # browser widget
        browserWidget = WebViewer(self,f'http://127.0.0.1:{port}')
        browserWidget.setSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
        
        # set browser as central widget
        self.setCentralWidget(browserWidget)
    
    def closeEvent(self, event):
        """
        captures the window close event [to be used later if required]
        """
        pass
        # print('The dash window received a close event')

##########################################
# when run on the commandline this code will be executed
#
if __name__ == "__main__":
       
    try:
        from docopt import docopt
    except ImportError:
        print('Install docopt using Anaconda:')
        print('    conda install -c anaconda docopt')
        print('or if not using Anaconda: ')
        print('    pip install docopt')
        print('or simply put the docopt.py script in the working folder')
        sys.exit(0)

    options = """dash-lineplot.py: Plotly dash line plotting utility.

        Usage:
          dash-lineplot.py [--configfile=<configFilename>] 
          dash-lineplot.py -h | --help 
 
        Options:
          -h, --help                           Show this screen.
          -f <configFilename>, --configfile <configFilename>    Excel config filename [default: ./dash-config.xlsx].
 
    """
    # process commandline arguments
    optionArguments = docopt(options)

    # always use callbacks
    # required for the slider, click data and rectangle tool to work
    useCallbacks = True
    # Excel file that defines the plots
    configfile = optionArguments["--configfile"]

    # extract the page title from the config file
    # read the config file
    cxls = pd.ExcelFile(configfile)
    dfPlotterHeader = pd.read_excel(cxls, 'header')
    dfPlotterHeader = dfPlotterHeader.set_index('Variable')
    pagetitle = dfPlotterHeader.loc['Pagetitle','Value'] if 'Pagetitle' in dfPlotterHeader.index else 'Dash flask server for plotting'

    # port used for the local Flask server
    port = '8050' 
           
    # start main app 
    appMain = QtWidgets.QApplication(sys.argv)
       
    # create new window and activate
    main_widget = DashPlotWindow(port,pagetitle)
    main_widget.show()

    # serve the required data to this window
    dashlineplotter = DashLinePlot()
    dashlineplotter.runPlotter(port, configfile, useCallbacks)
    
    # exit when main window closes
    sys.exit(appMain.exec_())
